import json
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from core.config import (
    MARKET_DATA_DB_PATH,
    NSE_CASH_DIR,
    NSE_FUTURES_DIR,
    NSE_OPTIONS_DIR,
    NSE_CHARTS_DIR,
)


def _parse_date(value):

    if isinstance(value, datetime):
        return value

    if value is None:
        return None

    for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(value), fmt)
        except ValueError:
            continue

    return None


def _pct_change(current, previous):

    if current is None or previous in (None, 0):
        return None

    return round(
        ((current - previous) / previous) * 100,
        2
    )


def _read_xlsx_rows(path):

    if not Path(path).exists():
        return []

    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True
    )

    try:
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)

        if not headers:
            return []

        headers = [str(h) for h in headers]

        return [
            dict(zip(headers, row))
            for row in rows
        ]

    finally:
        workbook.close()


class MarketContextRepository:

    def get_bse_cash_context(
        self,
        scrip_code,
        lookback=260
    ):

        if not Path(MARKET_DATA_DB_PATH).exists():
            return None

        conn = sqlite3.connect(
            MARKET_DATA_DB_PATH
        )

        cur = conn.cursor()

        cur.execute("""
        SELECT
            trade_date,
            fininstrm_id,
            symbol,
            open_price,
            high_price,
            low_price,
            close_price,
            prev_close,
            volume,
            turnover,
            trades
        FROM bse_cash
        WHERE fininstrm_id = ?
        """, (
            int(scrip_code),
        ))

        rows = cur.fetchall()

        conn.close()

        if not rows:
            return None

        rows = sorted(
            rows,
            key=lambda row: _parse_date(row[0]) or datetime.min
        )

        recent = rows[-lookback:]
        latest = recent[-1]
        previous = recent[-2] if len(recent) >= 2 else None

        volumes = [
            row[8]
            for row in recent[-21:-1]
            if row[8] is not None
        ]

        avg_volume_20 = (
            sum(volumes) / len(volumes)
            if volumes else None
        )

        volume_ratio = (
            round(latest[8] / avg_volume_20, 2)
            if latest[8] is not None
            and avg_volume_20
            else None
        )

        close_values = [
            row[6]
            for row in recent
            if row[6] is not None
        ]

        return {
            "source": "bse_cash_project",
            "trade_date": latest[0],
            "scrip_code": str(latest[1]),
            "symbol": latest[2],
            "open": latest[3],
            "high": latest[4],
            "low": latest[5],
            "close": latest[6],
            "prev_close": latest[7],
            "return_1d_pct": _pct_change(
                latest[6],
                previous[6] if previous else latest[7]
            ),
            "volume": latest[8],
            "avg_volume_20": round(avg_volume_20, 0)
            if avg_volume_20 else None,
            "volume_ratio_20": volume_ratio,
            "turnover": latest[9],
            "trades": latest[10],
            "high_52w": max(close_values) if close_values else None,
            "low_52w": min(close_values) if close_values else None,
        }

    def get_nse_cash_context(self, symbol):

        rows = _read_xlsx_rows(
            NSE_CASH_DIR / f"{symbol}.xlsx"
        )

        if not rows:
            return None

        rows = sorted(
            rows,
            key=lambda row: _parse_date(row.get("Date")) or datetime.min
        )

        latest = rows[-1]

        return {
            "source": "A_tech_indicator/Data/Cash",
            "trade_date": latest.get("Date"),
            "symbol": latest.get("Symbol"),
            "open": latest.get("Open"),
            "high": latest.get("High"),
            "low": latest.get("Low"),
            "close": latest.get("Close"),
            "prev_close": latest.get("Prev_Close"),
            "return_1d_pct": _pct_change(
                latest.get("Close"),
                latest.get("Prev_Close")
            ),
            "volume": latest.get("Traded_Volume"),
            "delivery_qty": latest.get("Deliverable_Qty"),
            "delivery_pct": latest.get("Delivery_%"),
            "series": latest.get("Series"),
        }

    def get_futures_context(self, symbol):

        rows = _read_xlsx_rows(
            NSE_FUTURES_DIR / f"{symbol}.xlsx"
        )

        if not rows:
            return None

        rows = sorted(
            rows,
            key=lambda row: _parse_date(row.get("Date")) or datetime.min
        )

        latest_date = rows[-1].get("Date")

        latest_rows = [
            row
            for row in rows
            if row.get("Date") == latest_date
        ]

        by_expiry = {}

        for row in latest_rows:
            expiry_type = row.get("Expiry_Type")

            by_expiry[expiry_type] = {
                "expiry": row.get("Expiry"),
                "close": row.get("Close_Price"),
                "net_change": row.get("Net_Change"),
                "traded_qty": row.get("Traded_Qty"),
                "open_interest": row.get("Open_Interest"),
                "change_in_oi": row.get("Change_in_OI"),
            }

        return {
            "source": "A_tech_indicator/Data/Futures",
            "trade_date": latest_date,
            "symbol": symbol,
            "expiries": by_expiry,
        }

    def get_options_context(self, symbol):

        rows = _read_xlsx_rows(
            NSE_OPTIONS_DIR / f"{symbol}.xlsx"
        )

        if not rows:
            return None

        rows = sorted(
            rows,
            key=lambda row: _parse_date(row.get("Date")) or datetime.min
        )

        latest_date = rows[-1].get("Date")

        latest_rows = [
            row
            for row in rows
            if row.get("Date") == latest_date
            and row.get("Expiry_Type") == "Current"
        ]

        ce_rows = [
            row for row in latest_rows
            if row.get("Option_Type") == "CE"
        ]

        pe_rows = [
            row for row in latest_rows
            if row.get("Option_Type") == "PE"
        ]

        ce_oi = sum(
            row.get("Open_Interest") or 0
            for row in ce_rows
        )

        pe_oi = sum(
            row.get("Open_Interest") or 0
            for row in pe_rows
        )

        return {
            "source": "A_tech_indicator/Data/Options",
            "trade_date": latest_date,
            "symbol": symbol,
            "current_expiry": (
                latest_rows[0].get("Expiry")
                if latest_rows else None
            ),
            "ce_open_interest": ce_oi,
            "pe_open_interest": pe_oi,
            "put_call_oi_ratio": round(pe_oi / ce_oi, 2)
            if ce_oi else None,
            "contracts_count": len(latest_rows),
        }

    def get_technical_context(self, symbol):

        path = NSE_CHARTS_DIR / f"{symbol}.json"

        if not path.exists():
            return None

        data = json.loads(
            path.read_text(
                encoding="utf-8"
            )
        )

        daily = (
            data
            .get("timeframes", {})
            .get("daily", {})
            .get("bars", [])
        )

        if not daily:
            return None

        latest = daily[-1]

        return {
            "source": "A_tech_indicator/Charts",
            "symbol": symbol,
            "last_updated": data.get("last_updated"),
            "date": latest.get("date"),
            "close": latest.get("close"),
            "rsi": latest.get("rsi"),
            "macd": latest.get("macd"),
            "macd_signal": latest.get("macd_signal"),
            "adx": latest.get("adx"),
            "supertrend": latest.get("supertrend"),
            "supertrend_dir": latest.get("supertrend_dir"),
            "sma20": latest.get("sma20"),
            "sma50": latest.get("sma50"),
            "sma100": latest.get("sma100"),
            "sma200": latest.get("sma200"),
            "volume_signal": latest.get("volume_signal"),
        }

    def get_context(
        self,
        scrip_code=None,
        symbol=None
    ):

        bse_cash = (
            self.get_bse_cash_context(scrip_code)
            if scrip_code else None
        )

        resolved_symbol = (
            symbol
            or (
                bse_cash.get("symbol")
                if bse_cash else None
            )
        )

        return {
            "scrip_code": str(scrip_code)
            if scrip_code else None,
            "symbol": resolved_symbol,
            "bse_cash": bse_cash,
            "nse_cash": self.get_nse_cash_context(resolved_symbol)
            if resolved_symbol else None,
            "futures": self.get_futures_context(resolved_symbol)
            if resolved_symbol else None,
            "options": self.get_options_context(resolved_symbol)
            if resolved_symbol else None,
            "technical": self.get_technical_context(resolved_symbol)
            if resolved_symbol else None,
        }
